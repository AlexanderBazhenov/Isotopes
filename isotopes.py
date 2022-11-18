import numpy as np
from mpmath import *
from twin import *
from intvalpy import *
from operation import *

import matplotlib.pyplot as plt
from openpyxl import load_workbook

workbook = load_workbook(filename="data\\table.xlsx")
sheet = workbook['Hydrogen']
'''поиск внешней границы по всем x(1H)'''
row_num = np.array([3,24])
col_num = [4, 5, 10, 11]

min_bound = list()
max_bound = list()

for j in col_num:
    atomic = list()
    for i in range(row_num[0], row_num[1]+1):
        atomic.append(sheet.cell(row=i, column=j).value)
    atomic = list(filter(None, atomic))

    if (j == col_num[0]) or (j == col_num[1]):
        min_bound.append(min(atomic))
    else:
        max_bound.append(max(atomic)) 

print(min_bound,max_bound)
mu1 = Interval(min_bound[0],max_bound[0])
mu2 = Interval(min_bound[1],max_bound[1])
mu = 1*mu1+2*mu2
print("внешняя граница: ",mu)

'''мода найдена внимательным взглядом'''
low = list()
low.append(sheet.cell(row=14, column=4).value)
low.append(sheet.cell(row=14, column=5).value)

up = list()
up.append(sheet.cell(row=16, column=10).value)
up.append(sheet.cell(row=16, column=11).value)

inter = 1*Interval(low[0],up[0])+2*Interval(low[1],up[1])

T = Twin(inter,mu)
print(T)

print(T.X_l)
print(T.X_l.a,T.X_l.b)

print(T.X)
print(T.X.a,T.X.b)

'''рисование твина. Не нашла готовой функции :('''
fig, ax = plt.subplots()
ax.plot( [T.X_l.a,T.X_l.b],[0,0],c = (0.0, 0.0, 1, 0.5),
        linewidth = 12)

ax.plot( [T.X.a,T.X.b],[0,0],c = (1, 0.0, 0.0, 0.5),
        linewidth = 12)

plt.show()
